"""Arma el libro de confirmación para Capital Humano."""

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(DIR, "confirmacion.json"), encoding="utf8") as f:
    D = json.load(f)

ARIAL = "Arial"
AZUL = "1F3864"
AMARILLO = PatternFill("solid", fgColor="FFF2CC")
GRIS = PatternFill("solid", fgColor="F2F2F2")
CAB = PatternFill("solid", fgColor="1F3864")
borde = Border(*[Side("thin", color="BFBFBF")] * 4)

wb = Workbook()


def cabecera(ws, cols, fila=1):
    for i, (titulo, ancho) in enumerate(cols, start=1):
        c = ws.cell(row=fila, column=i, value=titulo)
        c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        c.fill = CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[fila].height = 32
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


# ---------------------------------------------------------------- Instrucciones
ws = wb.active
ws.title = "Instrucciones"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 104

txt = [
    ("Confirmación del organigrama — CapitalIA", 16, True, AZUL),
    ("", 11, False, "000000"),
    (
        "Vamos a volcar en CapitalIA las 4.771 personas que liquidaron sueldo en julio de 2026, "
        "junto con el organigrama del sistema de sueldos (187 unidades). Casi todo engancha solo. "
        "Lo que está en las otras dos hojas es lo que NO pudimos resolver automáticamente y "
        "necesita que alguien que conoce el municipio lo confirme.",
        11,
        False,
        "000000",
    ),
    ("", 11, False, "000000"),
    ("Qué hay que hacer", 13, True, AZUL),
    (
        "Completar únicamente las celdas AMARILLAS. Todo lo demás es información para decidir "
        "y conviene no tocarlo.",
        11,
        False,
        "000000",
    ),
    ("", 11, False, "000000"),
    (
        "Hoja «Reparticiones» — 21 filas. Reparticiones que ya existen en CapitalIA y que el "
        "sistema de sueldos parece tener con otro nombre. La pregunta es si son la misma.",
        11,
        False,
        "000000",
    ),
    (
        "Hoja «Sectores» — 8 filas. Sectores donde hay gente liquidando y no está claro a qué "
        "repartición corresponden. Son 208 personas: si esto queda mal, esas personas terminan "
        "en la repartición equivocada.",
        11,
        False,
        "000000",
    ),
    ("", 11, False, "000000"),
    ("Por qué los nombres no coinciden", 13, True, AZUL),
    (
        "Los nombres del sistema de sueldos vienen truncados y con erratas —«SECRETARIA DE "
        "ORENAMIENTO Y CONV», «DESARROLO SUST», «COMUNIC.INSTITITUC»—. Los de CapitalIA están "
        "bien escritos. Al importar se conservan los de CapitalIA: del sistema de sueldos "
        "tomamos la estructura, no la redacción.",
        11,
        False,
        "000000",
    ),
    ("", 11, False, "000000"),
    ("Ejemplo de una fila completada", 13, True, AZUL),
]
r = 1
for t, size, bold, color in txt:
    c = ws.cell(row=r, column=2, value=t)
    c.font = Font(name=ARIAL, size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if size == 11 and len(t) > 90:
        ws.row_dimensions[r].height = 15 * (len(t) // 95 + 1)
    r += 1

ej_fila = r
ejemplo = [
    ("Repartición en CapitalIA", "DIR50 · Dirección de Arbolado Urbano"),
    ("Candidato en el sist. de sueldos", "DIRECCION DE ARBOLADO  (1.07.240)"),
    ("¿Es la misma? →", "SÍ"),
    ("Si no, ¿cuál es la correcta? →", ""),
    ("Observaciones →", "Le sacaron «Urbano» al abreviar"),
]
for etiqueta, valor in ejemplo:
    ws.cell(row=r, column=1, value=None)
    a = ws.cell(row=r, column=2, value=f"{etiqueta}   {valor}")
    a.font = Font(name=ARIAL, size=10, italic=True)
    a.fill = AMARILLO if "→" in etiqueta else GRIS
    a.border = borde
    r += 1

r += 1
c = ws.cell(row=r, column=2, value="Avance")
c.font = Font(name=ARIAL, size=13, bold=True, color=AZUL)
r += 1
for etiqueta, formula in [
    ("Reparticiones respondidas", '=COUNTIF(Reparticiones!I2:I22,"<>")&" de 21"'),
    ("Sectores respondidos", '=COUNTIF(Sectores!H2:H9,"<>")&" de 8"'),
]:
    ws.cell(row=r, column=2, value=etiqueta).font = Font(name=ARIAL, size=11)
    c = ws.cell(row=r, column=3, value=formula)
    c.font = Font(name=ARIAL, size=11, bold=True)
    r += 1
ws.column_dimensions["C"].width = 14


# ---------------------------------------------------------------- Reparticiones
ws = wb.create_sheet("Reparticiones")
cabecera(
    ws,
    [
        ("Código", 9),
        ("Repartición en CapitalIA", 42),
        ("Nivel", 13),
        ("Candidato en el sistema de sueldos", 42),
        ("Código", 11),
        ("Nivel", 13),
        ("Parecido", 9),
        ("Otra opción", 38),
        ("¿Es la misma?", 13),
        ("Si no, ¿cuál es la correcta?", 34),
        ("Observaciones", 34),
    ],
)

sino = DataValidation(type="list", formula1='"SÍ,NO,NO SÉ"', allow_blank=True)
ws.add_data_validation(sino)

for i, x in enumerate(D["reparticiones"], start=2):
    vals = [
        x["code"], x["nombre"], x["nivel"],
        x["candidato"] or "— sin candidato —", x["candidatoCode"], x["candidatoNivel"],
        x["parecido"], x["segunda"], None, None, None,
    ]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=ARIAL, size=10)
        c.border = borde
        c.alignment = Alignment(vertical="center", wrap_text=j in (2, 4, 8, 10, 11))
        if j >= 9:
            c.fill = AMARILLO
        if j == 7:
            c.number_format = "0%"
            c.alignment = Alignment(horizontal="center")
        if j == 4 and not x["candidato"]:
            c.font = Font(name=ARIAL, size=10, italic=True, color="C00000")
    sino.add(ws.cell(row=i, column=9))
    ws.row_dimensions[i].height = 30

ws.auto_filter.ref = f"A1:K{len(D['reparticiones']) + 1}"


# ---------------------------------------------------------------------- Sectores
ws = wb.create_sheet("Sectores")
cabecera(
    ws,
    [
        ("Personas", 10),
        ("Código", 9),
        ("Sector (como figura en la liquidación)", 42),
        ("Iría a esta repartición", 42),
        ("Código", 11),
        ("Parecido", 9),
        ("Otra opción", 42),
        ("¿Es correcto?", 13),
        ("Si no, ¿cuál es la correcta?", 34),
        ("Observaciones", 34),
    ],
)

sino2 = DataValidation(type="list", formula1='"SÍ,NO,NO SÉ"', allow_blank=True)
ws.add_data_validation(sino2)

for i, x in enumerate(D["sectores"], start=2):
    vals = [
        x["agentes"], x["codi"], x["sector"],
        x["candidato"] or "— ninguna se le parece —", x["candidatoCode"],
        x["parecido"], x["segunda"], None, None, None,
    ]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=ARIAL, size=10)
        c.border = borde
        c.alignment = Alignment(vertical="center", wrap_text=j in (3, 4, 7, 9, 10))
        if j >= 8:
            c.fill = AMARILLO
        if j == 1:
            c.font = Font(name=ARIAL, size=10, bold=True)
            c.alignment = Alignment(horizontal="center")
        if j == 6:
            c.number_format = "0%"
            c.alignment = Alignment(horizontal="center")
        if j == 4 and not x["candidato"]:
            c.font = Font(name=ARIAL, size=10, italic=True, color="C00000")
    sino2.add(ws.cell(row=i, column=8))
    ws.row_dimensions[i].height = 30

n = len(D["sectores"]) + 1
ws.auto_filter.ref = f"A1:J{n}"
c = ws.cell(row=n + 1, column=1, value=f"=SUM(A2:A{n})")
c.font = Font(name=ARIAL, size=10, bold=True)
c.border = borde
c.alignment = Alignment(horizontal="center")
c = ws.cell(row=n + 1, column=3, value="personas en juego en esta hoja")
c.font = Font(name=ARIAL, size=10, bold=True)

out = os.path.join(DIR, "Confirmacion-organigrama-CapitalIA.xlsx")
wb.save(out)
print(out)
